from rest_framework.views import APIView 
from rest_framework.response import Response
from rest_framework import status
from django.contrib.auth import authenticate, login
from .models import *
import openpyxl
from django.core.exceptions import ValidationError
from rest_framework.permissions import IsAuthenticated
from rest_framework_simplejwt.authentication import JWTAuthentication  
from datetime import datetime
from .serializers import *
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework import viewsets
from rest_framework.parsers import MultiPartParser, FormParser


# API đăng nhập
class ApiLoginView(APIView):
    # Không yêu cầu xác thực đối với API đăng nhập, vì api đăng nhập là điểm đầu tiên mà người dùng truy cập đến 
    # Người dùng chưa đăng nhập thì không có thông tin xác thực để gửi đến server
    authentication_classes = [] #quy định các lớp xác thực được sử dụng để xác thực người dùng
    permission_classes = [] # xác định các lớp phân quyền , người dùng chưa đăng nhập nên loại bỏ phân quyền 
    def post(self, request):
        # Kiểm tra nếu người dùng đã đăng nhập
        if request.user.is_authenticated:
            return Response({"error": "Bạn đã đăng nhập"}, status=status.HTTP_400_BAD_REQUEST)

        username = request.data.get('username')
        password = request.data.get('password')
        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            refresh = RefreshToken.for_user(user)
            access_token = str(refresh.access_token)
            refresh_token = str(refresh)
            return Response({
                "message": "Đăng nhập thành công",
                "access_token": access_token,  
                "refresh_token": refresh_token  
            }, status=status.HTTP_200_OK)
        else:
            return Response({"error": "Tên đăng nhập hoặc mật khẩu không chính xác"}, status=status.HTTP_401_UNAUTHORIZED)
        
#--------------------------------------API Đăng Ký-------------------------------------------------------
# Api đăng kí giáo viên
class APIRegisterTeacherView(APIView):
    authentication_classes = []  # JWTAuthentication 
    permission_classes = []  # IsAuthenticated, IsAdmin
    parser_classes = (MultiPartParser, FormParser)
    def post(self, request, *args, **kwargs):
        serializer = ExcelUploadSerializer(data=request.FILES)  
        if serializer.is_valid():
            file = serializer.validated_data['file']
            workbook = openpyxl.load_workbook(file)
            sheet = workbook["Danh sách giáo viên"]

            created_users = 0
            existing_users = 0

            for row in sheet.iter_rows(min_row=2, values_only=True):
                ma_dinh_danh = row[1]
                full_name = row[2]
                ngay_sinh = row[3]
                gioi_tinh = row[4]
                dan_toc = row[6]
                dien_thoai = row[7]
                vi_tri = row[8]  

                if not ma_dinh_danh or not full_name or not vi_tri:
                    continue

                is_teacher = vi_tri.lower() == 'giáo viên'
                is_admin = vi_tri.lower() == 'cán bộ quản lý'
                
                if not is_teacher and not is_admin:
                    continue  

                if CustomUser.objects.filter(user_id=ma_dinh_danh).exists():
                    existing_users += 1
                    continue

                if ngay_sinh:
                    try:
                        ngay_sinh = datetime.strptime(ngay_sinh, '%d/%m/%Y').date()
                    except ValueError:
                        continue  

                user = CustomUser(
                    username=ma_dinh_danh,  
                    user_id=ma_dinh_danh,
                    phone_number=dien_thoai,
                    is_teacher=is_teacher,
                    is_admin=is_admin,
                )
                try:
                    user.save()

                    if is_teacher:
                        Teacher.objects.create(
                            user=user,
                            full_name=full_name,
                            sex=gioi_tinh,
                            day_of_birth=ngay_sinh,
                            nation=dan_toc,
                            active_status=row[5], 
                            contract_types=row[10],  
                            expertise_levels=row[11], 
                            subjects=row[12] 
                        )
                    elif is_admin:
                        Admin.objects.create(
                            user=user,
                            full_name=full_name,
                            sex=gioi_tinh,
                            day_of_birth=ngay_sinh,
                            nation=dan_toc,
                            active_status=row[5],  
                            contract_types=row[10],  
                            expertise_levels=row[11],  
                            description=row[9]  
                        )
                    created_users += 1
                except ValidationError as e:
                    return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

            if created_users > 0:
                return Response({"detail": f"Tạo thành công {created_users} tài khoản giáo viên/cán bộ quản lý."}, status=status.HTTP_200_OK)
            else:
                return Response({"detail": "Không có tài khoản mới được tạo."}, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

# Api đăng ký học sinh
class APIRegisterStudentView(APIView):
    authentication_classes = []  # JWTAuthentication 
    permission_classes = []  # IsAuthenticated, IsAdmin
    parser_classes = (MultiPartParser, FormParser)
    def post(self, request, *args, **kwargs):
        serializer = ExcelUploadSerializer(data=request.data)
        if serializer.is_valid():
            file = serializer.validated_data['file']
            workbook = openpyxl.load_workbook(file)
            sheet = workbook["Danh sách học sinh"]

            created_students = 0
            existing_students = 0

            for row in sheet.iter_rows(min_row=7, values_only=True):
                ma_dinh_danh = row[2]
                full_name = row[3]
                ngay_sinh = row[4]
                gioi_tinh = row[5]
                dan_toc = row[6]
                active_status = row[7] 
                if not ma_dinh_danh or not full_name:
                    continue
                if CustomUser.objects.filter(user_id=ma_dinh_danh).exists():
                    existing_students += 1
                    continue
                if ngay_sinh:
                    try:
                        ngay_sinh = datetime.strptime(ngay_sinh, '%d/%m/%Y').date()
                    except ValueError:
                        continue  

                user = CustomUser(
                    username=ma_dinh_danh,  
                    user_id=ma_dinh_danh,
                    is_student=True,  
                )

                try:
                    user.save()
                    Student.objects.create(
                        user=user,
                        full_name=full_name,
                        sex=gioi_tinh,
                        nation=dan_toc,
                        day_of_birth=ngay_sinh,
                        parent=None, 
                        active_status = active_status  
                    )

                    created_students += 1
                except ValidationError as e:
                    return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

            if created_students > 0:
                return Response({"detail": f"Tạo thành công {created_students} tài khoản học sinh."}, status=status.HTTP_200_OK)
            else:
                return Response({"detail": "Không có tài khoản mới được tạo."}, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


# Dạng đăng kí bằng json 
# class APIRegisterStudentView(APIView):
#     authentication_classes = []  # JWTAuthentication 
#     permission_classes = []  # IsAuthenticated, IsAdmin
#     def post(self, request, *args, **kwargs):
#         students_data = request.data.get('students', [])  # Lấy danh sách học sinh từ JSON
#         if not students_data:
#             return Response({"detail": "Không có dữ liệu học sinh."}, status=status.HTTP_400_BAD_REQUEST)

#         created_students = 0
#         existing_students = 0

#         for student_data in students_data:
#             ma_dinh_danh = student_data.get('user_id')
#             full_name = student_data.get('full_name')
#             ngay_sinh = student_data.get('day_of_birth')
#             gioi_tinh = student_data.get('sex')
#             dan_toc = student_data.get('nation')
#             active_status = student_data.get('active_status')

#             if not ma_dinh_danh or not full_name:
#                 continue

#             # Kiểm tra xem học sinh đã tồn tại chưa
#             if CustomUser.objects.filter(user_id=ma_dinh_danh).exists():
#                 existing_students += 1
#                 continue

#             # Chuyển đổi ngày sinh từ string sang date
#             if ngay_sinh:
#                 try:
#                     ngay_sinh = datetime.strptime(ngay_sinh, '%Y-%m-%d').date()  # Format ngày sinh từ JSON
#                 except ValueError:
#                     continue  

#             # Tạo người dùng mới
#             user = CustomUser(
#                 username=ma_dinh_danh,  
#                 user_id=ma_dinh_danh,
#                 is_student=True,  
#             )

#             try:
#                 user.save()  # Lưu CustomUser
#                 # Tạo bản ghi Student
#                 Student.objects.create(
#                     user=user,
#                     full_name=full_name,
#                     sex=gioi_tinh,
#                     nation=dan_toc,
#                     day_of_birth=ngay_sinh,
#                     parent=None,  # Có thể xử lý parent nếu cần
#                     active_status=active_status
#                 )

#                 created_students += 1
#             except ValidationError as e:
#                 return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

#         if created_students > 0:
#             return Response({"detail": f"Tạo thành công {created_students} tài khoản học sinh."}, status=status.HTTP_200_OK)
#         else:
#             return Response({"detail": "Không có tài khoản mới được tạo."}, status=status.HTTP_200_OK)



class APIRegisterParentView(APIView):
    authentication_classes = []  # JWTAuthentication 
    permission_classes = []  # IsAuthenticated, IsAdmin
    parser_classes = (MultiPartParser, FormParser)
    def post(self, request, *args, **kwargs):
        serializer = ExcelUploadSerializer(data=request.FILES)  
        if serializer.is_valid():
            file = serializer.validated_data['file']
            workbook = openpyxl.load_workbook(file)
            sheet = workbook["Danh sách phụ huynh"]  

            created_parents = 0
            existing_parents = 0  
            invalid_students = 0  
            added_students_info = []
            added_students_messages = []
            for row in sheet.iter_rows(min_row=5, values_only=True):
                full_name = row[1]
                phone_number = row[2]
                address = row[3]
                ma_dinh_danh_hoc_sinh = row[4]

                if not full_name or not phone_number or not ma_dinh_danh_hoc_sinh:
                    continue

                try:
                    ma_dinh_danh_hoc_sinh = str(ma_dinh_danh_hoc_sinh).strip()
                    student = Student.objects.get(user__user_id=ma_dinh_danh_hoc_sinh)
                except Student.DoesNotExist:
                    invalid_students += 1
                    print(f"Học sinh với mã định danh {ma_dinh_danh_hoc_sinh} không tồn tại.")
                    continue
                parent_user = CustomUser.objects.filter(username=phone_number, is_parent=True).first()

                if parent_user:
                    if not student.parent:
                        student.parent = parent_user.parent
                        student.save()
                    existing_parents += 1
                    continue
                user = CustomUser(
                    username=phone_number,
                    user_id=phone_number,  
                    phone_number=phone_number,
                    is_parent=True,
                )
                user.set_password(phone_number)  
                user.save()
                parent = Parent.objects.create(
                    user=user,
                    full_name=full_name,
                    address=address
                )
                student.parent = parent
                student.save()
                created_parents += 1
                for info in added_students_info:
                    added_students_messages.append(f"Đã thêm học sinh {info['student_user_id']} cho phụ huynh {info['parent_user_id']}.")
            return Response({
                "detail": f"Tạo thành công {created_parents} tài khoản phụ huynh.",
                "existing_parents": existing_parents,
                "invalid_students": invalid_students,
                "added_students_messages": added_students_messages,
            }, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


# ---------------------------------------------------------------------------------------------
# api lấy thông tin user
class UserDetailView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        user = request.user
        user_data = UserSerializer(user).data
        
        # Xóa bỏ một số trường
        user_data.pop('password', None)
        user_data.pop('groups', None)
        user_data.pop('user_permissions', None)

        # Chỉ thêm vai trò nếu giá trị là True
        roles = ['is_teacher', 'is_admin', 'is_parent', 'is_student']
        for role in roles:
            if user_data[role]:  # chỉ thêm vào nếu là True
                user_data[role] = user_data[role]
            else:
                user_data.pop(role, None)  # loại bỏ nếu không phải
        # Lấy dữ liệu bổ sung từ các mô hình khác
        if user.is_teacher:
            try:
                teacher = user.teacher
                teacher_data = TeacherSerializer(teacher).data
                user_data.update(teacher_data)
            except Teacher.DoesNotExist:
                return Response({'error': 'Teacher profile not found.'}, status=status.HTTP_404_NOT_FOUND)
        elif user.is_admin:
            try:
                admin = user.admin
                admin_data = AdminSerializer(admin).data
                user_data.update(admin_data)
            except Admin.DoesNotExist:
                return Response({'error': 'Admin profile not found.'}, status=status.HTTP_404_NOT_FOUND)
        elif user.is_parent:
            try:
                parent = user.parent
                parent_data = ParentSerializer(parent).data
                user_data.update(parent_data)
            except Parent.DoesNotExist:
                return Response({'error': 'Parent profile not found.'}, status=status.HTTP_404_NOT_FOUND)
        elif user.is_student:
            try:
                student = user.student
                student_data = StudentSerializer(student).data
                user_data.update(student_data)
            except Student.DoesNotExist:
                return Response({'error': 'Student profile not found.'}, status=status.HTTP_404_NOT_FOUND)

        return Response(user_data, status=status.HTTP_200_OK)

# -----------------------------------------------view-set------------------------------------------

class CustomUserViewSet(viewsets.ModelViewSet):
    queryset = CustomUser.objects.all()
    serializer_class = UserSerializer
    lookup_field = 'user_id'

class TeacherViewSet(viewsets.ModelViewSet):
    queryset = Teacher.objects.all()
    serializer_class = TeacherSerializer
    lookup_field = 'user_id'

class AdminViewSet(viewsets.ModelViewSet):
    queryset = Admin.objects.all()
    serializer_class = AdminSerializer

class ParentViewSet(viewsets.ModelViewSet):
    queryset = Parent.objects.all()
    serializer_class = ParentSerializer

class StudentViewSet(viewsets.ModelViewSet):
    queryset = Student.objects.all()
    serializer_class = StudentSerializer

class CustomUserViewSet(viewsets.ModelViewSet):
    queryset = CustomUser.objects.all()
    serializer_class = UserUpdateSerializer
    permission_classes = []  
    def get_serializer_class(self):
        if self.action in ['partial_update', 'update']:
            return UserUpdateSerializer
        return super().get_serializer_class()

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)

        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)

        return Response(serializer.data, status=status.HTTP_200_OK)


# -----------------------------------api đổi mật khẩu----------------------------------------------
class ChangePasswordView(APIView):
    authentication_classes = [JWTAuthentication] 
    permission_classes = [IsAuthenticated]
    def post(self, request, *args, **kwargs):
        serializer = ChangePasswordSerializer(data=request.data, context={'request': request})
        if serializer.is_valid():
            user = request.user
            user.set_password(serializer.validated_data['new_password'])
            user.save()
            return Response({"detail": "Mật khẩu đã được thay đổi thành công."}, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

# -------------------------------------------------------------------------------------------

class AdminPasswordResetView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        if not request.user.is_admin:  # Kiểm tra quyền admin
            return Response({"detail": "Bạn không có quyền thực hiện thao tác này."}, status=status.HTTP_403_FORBIDDEN)

        serializer = AdminPasswordResetSerializer(data=request.data)
        if serializer.is_valid():
            user_id = serializer.validated_data['user_id']
            try:
                user = CustomUser.objects.get(user_id=user_id)
                new_password = user.user_id  # Reset mật khẩu về mã định danh
                user.set_password(new_password)
                user.save()
                return Response({"detail": f"Mật khẩu của người dùng {user.full_name} đã được reset thành công."}, status=status.HTTP_200_OK)
            except CustomUser.DoesNotExist:
                return Response({"error": "Người dùng không tồn tại."}, status=status.HTTP_404_NOT_FOUND)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
